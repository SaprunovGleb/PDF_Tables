# -*- coding: utf-8 -*-
"""
Created on Wed Jul 15 11:57:44 2020

@author: Flyin
"""
import openpyxl 
import sys
from testOPT import testOPT

def txtToXlsxOpt (pathInTable="PDFtoTXTTabula.txt",pathInTiket="PDFtoTXTpymupdf.txt", pathOut="1.xlsx", nameSourse="cmegroup",Nsdvig=500):
    #создаем файл Excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title ="1 sheet"
    #print(colLines)
    tiket =""
    rowSheet=1
    colSheet=1
    #создаем шапку и массив месяцев
    headSheet=["Tiket","Call/Put", "date",	"strike","GLOBEX OPEN",	"OPEN OUTCRY OPEN RANGE","GLOBEX HIGH","GLOBEX LOW",
               "OPEN OUTCRY HIGH",	"OPEN OUTCRY LOW", "OPEN OUTCRY CLOSE RANGE", "SETT.PRICE",
               "sign", "PT. CHGE","DELTA", "EXER CISES","OPEN OUTCRY VOLUME","GLOBEX®VOLUME", 
               "PNT VOLUME","OPEN INTEREST","sign","add"]
    months=["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]
    # записываем шапку
    for name in headSheet:
        sheet.cell(row=rowSheet, column=colSheet).value = name
        colSheet+=1
    rowSheet+=1
    wb.save(pathOut)
    #открываем файл таблица
    tableFile = open(pathInTable,"r")
    linesTabel = tableFile.readlines()
    colLinesTabel = len(linesTabel)
    nLineTabel = 0
    #открываем файл тикет
    tiketFile = open(pathInTiket,"r")
    linesTiket = tiketFile.readlines()
    colLinesTiket = len(linesTiket)
    nLineTiket = 1
    nLineTiketLast=0
    #создаем переменные
    lastStrike=-9999999
    strike=0
    day=999
    month = "DEC"
    year = 99999 
    dateLast = ""
    tiketLast=["-","-"]
    date="-"    
    nLineTiketLast=0
    nLineTiket = 1
    nLineTabel = 0
    #создаем массив который будем записывать в excel
    workLine=linesTabel[nLineTabel]
    workLine=workLine.replace("\"","")
    workLine=workLine.replace("*","")
    workLine=workLine.replace("#"," ")
    workLine=workLine.split(",")
    workMass =[]
    j=0
    workMass.append("")#место под тикет
    workMass.append("")#место под Opt, Call, Putt
    workMass.append("")#место под дату
    tiket =["",""]
    this=True

    while j<len(workLine):
        workMassBetween = workLine[j].split()
        k = 0
        while k < len(workMassBetween):
            workMass.append(workMassBetween[k].replace(".",","))
            k+=1                   
        j+=1
    #пербираем все строки из таблицы
    dig=0
    while ((nLineTabel+2)<colLinesTabel)and(dig<4):
        #print(workMass)

        if len(workMass)>9:
            look=True
        l=0
        while (nLineTiket<colLinesTiket)and look:
            masslinesTiket=linesTiket[nLineTiket].split()
            if ((linesTiket[nLineTiket].replace(" ",""))[0:-1]==workMass[3]):
                look = False
            elif len(masslinesTiket)>0:
                if masslinesTiket[0]==workMass[3]:
                    look = False
                else:
                    nLineTiket+=1     
                    l+=1
                if l>Nsdvig:
                    look = False
                    nLineTiket+=-l
            else:
                nLineTiket+=1
                l+=1
        #if nLineTabel == 220:
            #print(masslinesTiket[0],workMass[3])
        if (len(workMass[3]) >= 5)and(months.count(workMass[3][0:3])!=0)and(workMass[3][3:5].isdigit()):
            date =  workMass[3]
            #print(date)
            this=False
            look=True
            l=0
            while (nLineTiket<colLinesTiket)and look:
                masslinesTiket=linesTiket[nLineTiket].split()
                if ((linesTiket[nLineTiket].replace(" ",""))[0:-1]==workMass[3]):
                    look = False
                elif len(masslinesTiket)>0:
                    if masslinesTiket[0]==workMass[3]:
                        look = False
                    else:
                        nLineTiket+=1
                        l+=1
                    if l>Nsdvig:
                        look = False
                        nLineTiket+=-l
                        
                else:
                    nLineTiket+=1
            if (int(workMass[3][3:5])<year)or((int(workMass[3][3:5])==year) 
                        and (months.index(workMass[3][0:3])<months.index(month))):
                tiket =["",""]
            month=date[0:3]
            year=int(date[3:5])
            day=99
        
        elif (len(workMass[3]) >= 7)and((months.count(workMass[3][2:5])!=0)and workMass[3][5:7].isdigit()and(workMass[3][0:2].isdigit())):
            date =  workMass[3]
            #print(date)
            this=False
            look=True
            l=0
            while (nLineTiket<colLinesTiket)and look:
                masslinesTiket=linesTiket[nLineTiket].split()
                if ((linesTiket[nLineTiket].replace(" ",""))[0:-1]==workMass[3]):
                    look = False
                elif len(masslinesTiket)>0:
                    if masslinesTiket[0]==workMass[3]:
                        look = False
                    else:
                        nLineTiket+=1
                        l+=1
                    if l>Nsdvig:
                        look = False
                        nLineTiket+=-l
                        
                else:
                    nLineTiket+=1
            if (int(workMass[3][5:7])<year)or((int(workMass[3][5:7])==year) 
                        and ((months.index(workMass[3][2:5])<months.index(month))or((
                            months.index(workMass[3][2:5])==months.index(month))and int(workMass[3][0:2])<day))):
                tiket =["",""]
            month=date[2:5]
            year=int(date[5:7])    
            day=int(date[0:2])
            
        if len(workMass)>5 :
            if workMass[3]=="-":
                strike=int(workMass[4])
                strike=0-strike
                #print("1  ",strike)
            elif (workMass[3].isdigit()):
                strike=int(workMass[3])
                #print("2  ",strike)
            #print (dateLast,date)
            if strike<=lastStrike:
                date="-"
            
            lastStrike=strike

        if date == "-":
            #print("поиск даты")
            #print(nLineTiket,workMass[3],workMass[4])
            look=True
            l=0
            while (nLineTiket<colLinesTiket)and look:
                masslinesTiket=linesTiket[nLineTiket].split()
                if ((linesTiket[nLineTiket].replace(" ",""))[0:-1]==workMass[3]):
                    look = False
                elif len(masslinesTiket)>0:
                    if masslinesTiket[0]==workMass[3]:
                        look = False
                    else:
                        nLineTiket+=1 
                        l+=1
                    if l>Nsdvig:
                        look = False
                        nLineTiket+=-l
                        
                else:
                    nLineTiket+=1
                    l+=1
            dopLineTiket = 0
            con=True
            while con:
                #print (nLineTiket-dopLineTiket)
                if ((nLineTiket-dopLineTiket)==0):
                    con=False
                    print("дата не найдена")
                elif  (len(linesTiket[nLineTiket-dopLineTiket])>4  and(months.count(linesTiket[nLineTiket-dopLineTiket][0:3])!=0)and(
                    ((linesTiket[nLineTiket-dopLineTiket][3:5]).isdigit()))):
                    con=False
                    date=linesTiket[nLineTiket-dopLineTiket][0:5]
                    #print(date)
                    month=date[0:3]
                    year=int(date[3:5]) 
                    day=99
                elif (len(linesTiket[nLineTiket-dopLineTiket])>6  and(months.count(linesTiket[nLineTiket-dopLineTiket][2:5])!=0)and(
                    ((linesTiket[nLineTiket-dopLineTiket][5:7]).isdigit()))and(
                    ((linesTiket[nLineTiket-dopLineTiket][0:2]).isdigit()))):
                    con=False
                    date=linesTiket[nLineTiket-dopLineTiket][0:7]
                    #print(date)
                    month=date[2:5]
                    year=int(date[5:7]) 
                    day=int(date[0:2])
                else:
                    dopLineTiket+=1
            
            tiket =["",""]
            #if nLineTabel == 4932:
                #print(workMass, date, tiket)
        #print(date)
        #input()
        workMass[2]=date
        #print (nLineTiket,dopLineTiket,nLineTabel, )
        #print (linesTiket[nLineTiket], linesTabel[nLineTabel] )
        if tiket[0] == "":
            dopLineTiket = 0
            
            tiketProb=linesTiket[nLineTiket-dopLineTiket].split()
            con=True
            look = True
            l=0
            while (nLineTiket<colLinesTiket)and look:
                masslinesTiket=linesTiket[nLineTiket].split()
                if ((linesTiket[nLineTiket].replace(" ",""))[0:-1]==workMass[3]):
                    look = False

                elif len(masslinesTiket)>0:
                    if masslinesTiket[0]==workMass[3]:
                        look = False
                    else:
                        nLineTiket+=1
                        l+=1
                    if l>Nsdvig:
                        look = False
                        nLineTiket+=-l
                        
                else:
                    nLineTiket+=1
                    l+=1
            #print (nLineTiket,dopLineTiket,nLineTabel)
            while con:
                if ((nLineTiket-dopLineTiket)==1):
                    con=False
                    print("Тикет не найден")
                if len(tiketProb)==2:
                    if ((tiketProb[1]=="OPT")or(tiketProb[1]=="PUT")or(tiketProb[1]=="CALL")or(tiketProb[1]=="OFF")):
                        if nLineTiketLast==nLineTiket and linesTiket[nLineTiket-dopLineTiket-1]!="VOLUME\n":
                            #print (tiketLast,tiketProb,tiket)
                            #print("ищем вниз")
                            con2=True
                            dopLineTiket=0
                            nLineTiket+=1
                            tiketProb=linesTiket[nLineTiket+dopLineTiket].split()
                            
                            dopLineTiket+=1                          
 
                            while con2:
                                if (nLineTiket+1>colLinesTiket):
                                    con2=False
                                    print("Тикет не найден")
                                if len(tiketProb)==2:
                                    if (((tiketProb[1]=="OPT")or(tiketProb[1]=="PUT")or(tiketProb[1]=="CALL")
                                         or(tiketProb[1]=="OFF")))and nLineTiketLast!=nLineTiket:
                                        nLineTiketLast=nLineTiket
                                        #print(tiketProb)
                                        break                                        
                                if nLineTiket>(nLineTabel*30+100):
                                    sys.exit()
                                tiketProb=linesTiket[nLineTiket+dopLineTiket].split()
                                ntiketProb=nLineTiket+dopLineTiket
                                dopLineTiket+=1                          
                        #print (tiketLast,tiketProb,tiket)
                        
                        tiket=tiketProb
                        nLineTiketLast=ntiketProb
                        rowSheet+=1
                        con=False
                        nLineTiket+=1
                # if nLineTiket>(nLineTabel*20+100):
                #     sys.exit()
                #print (nLineTiket,dopLineTiket,nLineTabel, )
                tiketProb=linesTiket[nLineTiket-dopLineTiket].split()
                ntiketProb=nLineTiket-dopLineTiket
                dopLineTiket+=1
        #while ((nLineTiket+1)<colLinesTiket)and((linesTiket[nLineTiket].replace(" ",""))[0:-1]!=workMass[3]):
            #nLineTiket+=1
        workMass[0]=str(tiket[0])
        workMass[1]=str(tiket[1])
        #print(nLineTabel ,nLineTiket,tiket, date,workMass)
        #print("\n")
        if (len(workMass[3]) >= 5)and(workMass[3][0:5]=="TOTAL"):

            #print(nLineTabel,nLineTiket,dopLineTiket,tiket, date,workMass)
            this=False
            date = "-"

        #запись рабочего масива в файл
        if len(workMass)<9:
            this=False
        if workMass[3] ==workMass[0]:
            this=False
            dig+=1
            rowSheet+=-1
        if this:
            
# =============================================================================
#             while (nLineTiket<colLinesTiket)and look:
#                 masslinesTiket=linesTiket[nLineTiket].split()
#                 if ((linesTiket[nLineTiket].replace(" ",""))[0:-1]==workMass[3]):
#                     look = False
# 
#                 elif len(masslinesTiket)>0:
#                     if masslinesTiket[0]==workMass[3]:
#                         look = False
#                     else:
#                         nLineTiket+=1
#                 else:
#                     nLineTiket+=1
# =============================================================================
                    
            j=0
            d=1         
            #input()

            while j<len(workMass):
                if j==3:
                    if (workMass[j]=="-")or(workMass[j]=="+"):
                        d+=-1
                        j+=1
                        workMass[j] = str(workMass[j-1]) + str(workMass[j])
                elif (j+d)==13:
                    if ((workMass[j][0]=="-")or(workMass[j][0]=="+"))and workMass[j][-1].isdigit and (workMass[j][-1]!="-")and(workMass[j][-1]!="+"):
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j][0]
                        workMass[j]=workMass[j][1:]
                        d+=1
                    if (workMass[j]=="NEW")or(workMass[j]=="UNCH"):
                        d+=1
                elif (j+d == 7)or(j+d == 9):
                    workMassBetween = workMass[j].split("/")
                    if len(workMassBetween)==2:
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMassBetween[0]
                        d+=1
                        workMass[j]=workMassBetween[1]
                    elif workMass[j] == "----":
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j]
                        d+=1
                elif (workMass[j]=="NEW")or(workMass[j]=="UNCH"):
                    d+=1
                if (j+d)==10:
                    if workMass[j][-1]=="-"and (workMass[j][0].isdigit)and (workMass[j][0]!="-"):
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j].replace("-","")                        
                        workMass[j]="----"
                        d+=1
                sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j]
                j+=1
            dateLast=workMass[2]    
            sheet.cell(row=rowSheet, column=23).value = nLineTabel
            sheet.cell(row=rowSheet, column=24).value = nLineTiket
            nLineTiket+=1
            rowSheet+=1
            if (j+d)>8:
                dig=0
            else:
                dig+=1
        ScanNewLine = True
        while ScanNewLine and ((nLineTabel+2)<colLinesTabel):
            nLineTabel+=1    
            workLine=linesTabel[nLineTabel]
            workLine=workLine.replace("*","")
            workLine=workLine.replace("\"","")
            workLine=workLine.replace("#"," ")
    
            workLine=workLine.split(",")
            workMass =[]
            j=0
            workMass.append("")#место под тикет
            workMass.append("")#место под Opt, Call, Putt
            workMass.append("")#место под дату
            this=True
            while j<len(workLine):
                workMassBetween = workLine[j].split()
                k = 0
                while k < len(workMassBetween):
                    workMass.append(workMassBetween[k].replace(".",","))
                    k+=1                   
                j+=1
            ScanNewLine = False
            if len (workMass)>5:
                if workMass[4]=="OPT" or workMass[4]=="PUT" or workMass[4]=="OFF" or workMass[4]=="CALL":
                    ScanNewLine = True
                if workMass[3]=="-":
                    workMass[4]=str(workMass[3])+str(workMass[4])
                    workMass.pop(3)
            
            #if nLineTabel == 4932:
                #print(workMass, date, tiket)
                
                #wb.save(pathOut)
    wb.save(pathOut)
    pathErOut=pathOut[:-5]+"er.xlsx"
    if testOPT(pathInTable=pathOut,pathOut=pathErOut) == False:
        txtToXlsxOpt(pathInTable,pathInTiket, pathOut, nameSourse, Nsdvig=Nsdvig+300)
    
    
    
def txtToXlsxFut (pathInTable="PDFtoTXTTabula.txt",pathInTiket="PDFtoTXTpymupdf.txt", pathOut="1.xlsx", nameSourse="cmegroup"):
    #создаем файл Excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title ="1 sheet"
    #print(colLines)
    tiket =""
    rowSheet=1
    colSheet=1
    #создаем шапку и массив месяцев
    headSheet=["Tiket",	"date",	"open",	"high","low","sett. price",	"sign",
               "PT. CHGE","GLOBEX®VOLUME", "PNT/PIT VOLUME","OPEN INTEREST", "sign","add"]
    months=["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]
    # записываем шапку
    for name in headSheet:
        sheet.cell(row=rowSheet, column=colSheet).value = name
        colSheet+=1
    rowSheet+=1
    wb.save(pathOut)
    #открываем файл таблица
    tableFile = open(pathInTable,"r")
    linesTabel = tableFile.readlines()
    colLinesTabel = len(linesTabel)
    nLineTabel = 0
    #открываем файл тикет
    tiketFile = open(pathInTiket,"r")
    linesTiket = tiketFile.readlines()
    colLinesTiket = len(linesTiket)
    #создаем переменные
    nLineTiket = 1
    date = 999
    year = 99999 
    month = "JAN"
    nLineTiketLast=-1
    #Работаем в цикле, пока не выйдем за границы файла или не дойдем до разделяющей линии
    while (nLineTabel<colLinesTabel)and(linesTabel[nLineTabel][0:8]!="DELIVERY")and(linesTabel[nLineTabel][0:8]!="SETTLEDT"):
        #условие для нахождения строк начинающихся с ddmmmyy(если не начинаются пропускаем строку) 
        if (not(linesTabel[nLineTabel][5:7].isdigit()))or(months.count(linesTabel[nLineTabel][2:5])==0)or(not(linesTabel[nLineTabel][5:7].isdigit())):
            nLineTabel+=1
        else:
            #если новая дата меньше старой, то нужно искать новый тикет
            if (int(linesTabel[nLineTabel][5:7])<year)or((int(linesTabel[nLineTabel][5:7])==year) 
                        and (months.index(linesTabel[nLineTabel][2:5])<months.index(month))) or (
                        (months.index(linesTabel[nLineTabel][2:5])==months.index(month))and (int(linesTabel[nLineTabel][0:2])<=date)):
                #print(nLineTabel,linesTabel[nLineTabel][0:5])   
                #спускаемся по файлу тикетов до нахждения такой-же даты
                while (nLineTiket<colLinesTiket)and(linesTiket[nLineTiket][0:7]!=linesTabel[nLineTabel][0:7]):
                    nLineTiket+=1
                #ищем на одну строчку выше тикет(он состоит из 2 частей имени и FUT)
                tiket = linesTiket[nLineTiket-1][0:-1]
                nLineTiketLast=nLineTiket
                #print(nLineTiket,linesTiket[nLineTiket][0:5])
                rowSheet+=1
                pTiket=tiket.split()
                #print(pTiket)
                #если не нашли тикет на строчке выше ищем ниже TOTAL + тикет
                if len(pTiket)!=2 or pTiket[1]!="FUT":
                    #print("поиск вниз")
                    nLineTiketDop=nLineTiket
                    while (nLineTiket<colLinesTiket)and(linesTiket[nLineTiketDop][0:5]!="TOTAL"):
                        nLineTiketDop+=1
                    tiket = linesTiket[nLineTiketDop][7:-1]
            #спускаемся по файлу тикетов до нахждения такой-же даты
            while (nLineTiket<colLinesTiket)and(linesTiket[nLineTiket][0:7]!=linesTabel[nLineTabel][0:7]):
                nLineTiket+=1
            #спускаемся по файлу тикетов на ещё одну строчку чтобы не было перекрытия при одинаковых датах
            nLineTiket+=1
            #записываем в отдельные переменные новые день, месяц, год
            date = int(linesTabel[nLineTabel][0:2])
            month=linesTabel[nLineTabel][2:5]
            year=int(linesTabel[nLineTabel][5:7])
        #создаем рабочий массив из строки из файла таблиц
            #считываем строчку
            workLine=linesTabel[nLineTabel]
            #делим на слова используя как разделитель (",")
            workLine=workLine.split(",")
            #создаем массив
            workMass =[]
            #задаем номер слова из workLine, работаем в цикле пока j< длины массива
            j=0
            while j<len(workLine):
                #если слово не пустое
                if workLine[j] !="":
                    #разделяем на слова используя как разделитель (" ")
                    workMassBetween = workLine[j].split()
                    #перебираем все слова из промежуточного массива и добавляем в рабочий
                    k = 0
                    while k < len(workMassBetween):
                        #при добавлении заменяем . на , для того, чтоб excel определял числа
                        workMass.append(workMassBetween[k].replace(".",","))
                        k+=1            
                #переходим на следующее слово из workLine
                j+=1
            #записываем тикет в эксель файл
            sheet.cell(row=rowSheet, column=1).value = tiket
            #month=workMass[0][0:3]
            #year=int(workMass[0][3:5])
            #счетчик слова из массива и сдвиг в колонках
            j=0
            d=2
            # проходим по всем словам из рабочего массива
            while j<len(workMass):
                # high/low могут быть ----, "цифры" и "/цифры", "цифры/цифры"
                if j == 2:
                    workMassBetween = workMass[j].split("/")
                    sheet.cell(row=rowSheet, column=(j+d)).value = workMassBetween[0]
                    if len(workMassBetween) == 2:
                        d+=1
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMassBetween[1]
                    if workMass[j][1]=="-":
                        d+=1
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j]
                
                elif j==3:
                    if workMass[j][0]=="/":
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j][1:]
                    else:
                        if d == 2:
                            d+=1
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j]
                #UNCH и NEW требуют пропуска столбца перед их записью
                elif (workMass[j] == 'UNCH')or(workMass[j] == 'NEW'):
                    d+=1
                    sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j]
                #в 7 столбик записывается знак, но он бывает находится в отдельном слове, а бывает нет, необходима проверка
                elif (j+d)==7:
                    if (len(workMass[j])!=1)and((workMass[j][0]=="-")or(workMass[j][0]=="+")):
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j][0]
                        d+=1
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j][1:]
                    else:
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j]
                else:
                    sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j]
                j+=1
            sheet.cell(row=rowSheet, column=15).value = nLineTabel
            sheet.cell(row=rowSheet, column=16).value = nLineTiket
            #if nLineTabel%200 == 0:
                #print(workMass)
                #wb.save(pathOut)
            rowSheet+=1
            nLineTabel+=1
        #if nLineTabel>100:
            #nLineTabel=colLinesTabel    
    nLineTiket = 1
    nLineTabel = 0
    date = 999
    year = 99999 
    month = "JAN"

    rowSheet+=1
    while (nLineTabel<colLinesTabel)and(linesTabel[nLineTabel][0:8]!="DELIVERY")and(linesTabel[nLineTabel][0:8]!="SETTLEDT"):
        if (months.count(linesTabel[nLineTabel][0:3])==0)or(not(linesTabel[nLineTabel][3:5].isdigit())):
            nLineTabel+=1
        else:
            if (int(linesTabel[nLineTabel][3:5])<year)or((int(linesTabel[nLineTabel][3:5])==year) 
                        and (months.index(linesTabel[nLineTabel][0:3])<=months.index(month))):
                #print(nLineTabel,linesTabel[nLineTabel][0:5])                        
                while (nLineTiket<colLinesTiket)and(linesTiket[nLineTiket][0:5]!=linesTabel[nLineTabel][0:5]):
                    nLineTiket+=1
                tiket = linesTiket[nLineTiket-1][0:-1]
                
                #print(nLineTiket,linesTiket[nLineTiket][0:5])
                rowSheet+=1
                pTiket=tiket.split()
                #print(pTiket)
                #если не нашли тикет на строчке выше ищем ниже TOTAL + тикет
                if len(pTiket)!=2 or pTiket[1]!="FUT":
                    nLineTiketDop=nLineTiket
                    while (nLineTiket<colLinesTiket)and(linesTiket[nLineTiketDop][0:5]!="TOTAL"):
                        nLineTiketDop+=1
                    tiket = linesTiket[nLineTiketDop][7:-1]
                if tiket=="  ":
                    break
            while (nLineTiket<colLinesTiket)and(linesTiket[nLineTiket][0:5]!=linesTabel[nLineTabel][0:5]):
                nLineTiket+=1
            nLineTiket+=1
            month=linesTabel[nLineTabel][0:3]
            year=int(linesTabel[nLineTabel][3:5])    
            workLine=linesTabel[nLineTabel]
            workLine=workLine.split(",")
            workMass =[]
            j=0
            while j<len(workLine):
                if workLine[j] !="":
                    workMassBetween = workLine[j].split()
                    k = 0
                    while k < len(workMassBetween):
                        workMass.append(workMassBetween[k].replace(".",","))
                        k+=1                   
                j+=1
            sheet.cell(row=rowSheet, column=1).value = tiket
            month=workMass[0][0:3]
            year=int(workMass[0][3:5])
            j=0
            d=2
            
            while j<len(workMass):
            
                if j == 2:
                    workMassBetween = workMass[j].split("/")
                    sheet.cell(row=rowSheet, column=(j+d)).value = workMassBetween[0]
                    if len(workMassBetween) == 2:
                        d+=1
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMassBetween[1]
                    if workMass[j][1]=="-":
                        d+=1
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j]
                
                elif j==3:
                    if workMass[j][0]=="/":
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j][1:]
                    else:
                        if d == 2:
                            d+=1
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j]
                elif (workMass[j] == 'UNCH')or(workMass[j] == 'NEW'):
                    d+=1
                    sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j]
                elif (j+d)==7:
                    if (len(workMass[j])!=1)and((workMass[j][0]=="-")or(workMass[j][0]=="+")):
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j][0]
                        d+=1
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j][1:]
                    else:
                        sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j]
                else:
                    sheet.cell(row=rowSheet, column=(j+d)).value = workMass[j]
                j+=1
            sheet.cell(row=rowSheet, column=15).value = nLineTabel
            sheet.cell(row=rowSheet, column=16).value = nLineTiket
            #if nLineTabel%200 == 0:
                #print(workMass)
                #wb.save(pathOut)
            rowSheet+=1
            nLineTabel+=1
        #if nLineTabel>100:
            #nLineTabel=colLinesTabel
    tableFile.close()
    tiketFile.close()
    wb.save(pathOut)
              
#txtToXlsxOpt()     
#txtToXlsxFut()    
#                     sheet.cell(row=(j+1), column=(k+1)).value = table[k][i]