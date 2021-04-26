# -*- coding: utf-8 -*-
"""
Created on Sat Jul 18 17:08:47 2020

@author: Flyin
"""


import openpyxl 
wb = openpyxl.load_workbook("1.xlsx")
j=1
sheet = wb.active
print(sheet.cell(row=2, column=1).value)

while (sheet.cell(row=j, column=1).value !=None)or(sheet.cell(row=j+1, column=1).value !=None):
    while (sheet.cell(row=j, column=1).value !=None):
        j+=1
    print(j," Корректно",sheet.cell(row=j+1, column=1).value)
    if sheet.cell(row=j+1, column=1).value == sheet.cell(row=j-1, column=1).value:
        print(j," Сломано")
        break
    
    j+=1

