# -*- coding: utf-8 -*-
"""
Created on Fri Jul 24 14:02:27 2020

@author: Flyin
"""

import openpyxl 

def testOPT(pathInTable="1.xlsx",pathInTiket="PDFtoTXTpymupdf.txt", pathOut="test1.xlsx"):
    wbOut = openpyxl.Workbook()
    sheetOut = wbOut.active
    sheetOut.title ="1 sheet"
    #print(colLines)
    rowSheetOut=1
    tiketLast=["",""]
    #открываем файл тикет
    tiketFile = open(pathInTiket,"r")
    linesTiket = tiketFile.readlines()
    colLinesTiket = len(linesTiket)
    nLineTiket = 0
    while nLineTiket<colLinesTiket and linesTiket[nLineTiket]!="OPTIONS EOO'S AND BLOCKS\n":
        tiketProb=linesTiket[nLineTiket].split()
        if len(tiketProb)==2:
            if ((tiketProb[1]=="OPT")or(tiketProb[1]=="PUT")or(tiketProb[1]=="CALL")or(tiketProb[1]=="OFF")):
                if (tiketProb[0] == tiketLast[0])and(tiketProb[1] == tiketLast[1]):
                    nLineTiket+=1
                else: 
                    sheetOut.cell(row=rowSheetOut, column=1).value=str(tiketProb[0]+" "+tiketProb[1])
                    sheetOut.cell(row=rowSheetOut, column=2).value=str(tiketProb[0]+" "+tiketProb[1])
                    tiketLast=tiketProb
                    rowSheetOut+=1
        nLineTiket+=1
    wbIn = openpyxl.load_workbook(pathInTable, read_only=False)
    sheetIN = wbIn.worksheets[0]
    rowSheetOut=1
    #wbOut.save(pathOut)
    #input()
    for row in sheetIN.rows:
        tiket=str(row[0].value)+" "+str(row[1].value)
        if tiket!="None None":
            #print(tiket)
            #print(sheetOut.cell(row=rowSheetOut, column=1).value)
            if tiket!="Tiket Call/Put":
                while tiket!= sheetOut.cell(row=rowSheetOut, column=1).value:
                    rowSheetOut+=1
                sheetOut.cell(row=rowSheetOut, column=2).value="Exist"
    #wbOut.save(pathOut)
    #input()

    massErorr=[]
    for row in sheetOut.rows:
        if row[1].value!="Exist":
            massErorr.append(row[1].value)
    print("Ошибки:", massErorr)
    if len(massErorr)>0 and len(massErorr)<10:
        wbOut = openpyxl.Workbook()
        sheetOut = wbOut.active
        sheetOut.title ="1 sheet"
        #print(massErorr) 
        j=0
        sheetOut.cell(row=1, column=1).value=("Не занесённые тикеты за день")
        while j< len(massErorr):
            sheetOut.cell(row=j+2, column=1).value=massErorr[j]
            j+=1
        wbOut.save(pathOut)
        return True
    else: 
        if len(massErorr)>9:
            return False
        if len(massErorr)==0:
            return True
#testOPT()